import sys
import openpyxl
 
from docxtpl import DocxTemplate
from io import TextIOBase, TextIOWrapper
from PyQt5 import QtWidgets
 

SETTINGS = {}
 

class TemplateWordGenerator(QtWidgets.QDialog):
    def __init__(self):
        super().__init__()
        self.template_docx = None
        self.context = {}
        self.resize(420, 240)
        self.setWindowTitle('Шаблонизатор Microsoft-Word')
        self.button_for_getting_a_template = QtWidgets.QPushButton('Выберите файл шаблона', self)
        self.button_for_getting_a_template.move(120, 20)
        self.button_for_getting_a_template.clicked.connect(self.getting_a_template)
        self.button_for_getting_context_for_insertion = QtWidgets.QPushButton('Выберите файл, откуда\nбудут браться значения', self)
        self.button_for_getting_context_for_insertion.move(120, 60)
        self.button_for_getting_context_for_insertion.clicked.connect(self.getting_context_for_insertion)
        self.button_for_save_new_file = QtWidgets.QPushButton('Введите куда надо\nсохранить файл', self)
        self.button_for_save_new_file.move(120, 115)
        self.button_for_save_new_file.clicked.connect(self.save_new_file)
        self.button_for_directory = QtWidgets.QPushButton('Выберите папку, где лежит\n"шаблонизация файлов.xlsx"', self)
        self.button_for_directory.move(120, 170)
        self.button_for_directory.clicked.connect(self.getting_for_directory)


    def error_message(self, message: str):
        send_message = QtWidgets.QMessageBox()
        send_message.setIcon(QtWidgets.QMessageBox.Critical)
        send_message.setWindowTitle('Внимание!')
        send_message.setText('Ошибка')
        send_message.setInformativeText(message)
        send_message.exec_()
    
 
    def complete_message(self, message: str):
        send_message = QtWidgets.QMessageBox()
        send_message.setIcon(QtWidgets.QMessageBox.Information)
        send_message.setWindowTitle('Завершение записи')
        send_message.setInformativeText(message)
        send_message.exec_()
 

    def download_dict_settings_from_excel(self, path: str) -> dict:
        wb = openpyxl.load_workbook(f'{path}\\шаблонизация файлов.xlsx')
        worksheet = wb['шаблонизация']
        for col in worksheet.iter_cols(1, worksheet.max_column):
            for row in range(0, 1):
                SETTINGS[col[row].value] = []
            for rows in range(1, worksheet.max_row):
                SETTINGS[col[row].value].append(col[rows].value)
        return SETTINGS
 

    def getting_a_template(self, path: str=None, filename_template: str=None) -> TextIOWrapper:
        try:
            # укажем файл(путь до файла, вместе с его названием)
            if filename_template == None:
                path = QtWidgets.QFileDialog.getOpenFileName(
                    parent=self,
                    caption='Выберите файл шаблона',
                    filter='DOCX File (*.docx)'
                )
                template_file = (((path[0].replace('/', '\\')).strip()))
            else:
                template_file = f'{path}{filename_template}.docx'
            # получим объект файла
            self.template_docx = DocxTemplate(template_file=template_file)
            return self.template_docx
        except BaseException:
            self.error_message(message='Необходимо указать\n"шаблон файла.docx"')
 

    def getting_context_for_insertion(self, path: str=None, filename_replacement: str=None) -> dict:
        try:
            # укажем файл(путь до файла, вместе с его названием)
            if filename_replacement == None:
                path = QtWidgets.QFileDialog.getOpenFileName(
                    parent=self,
                    caption='Выберите файл, откуда\nбудут браться значения',
                    filter='TXT File (*.txt)'
                )
                filepath_txt = ((path[0].replace('/', '\\')).strip())
                file_txt = open(file=filepath_txt, mode='r', encoding='utf-8')
            else:
                file_txt = open(
                    file=f'{path}{filename_replacement}.txt',
                    mode='r',
                    encoding='utf-8'
                )
            while True:
                # считываем строку
                line = file_txt.readline().strip()
                # прерываем цикл, если строка пустая
                if not line:
                    break
                # разбиваем строку - "ключ=значение"
                key_value_for_context = line.split('=')
                # создаем словарь в виде "ключ: значение"
                self.context[(key_value_for_context[0]).strip()] = (key_value_for_context[1]).strip()
            # закрываем файл
            file_txt.close
            return self.context
        except BaseException:
            self.error_message(message='Необходимо указать\n"Файл значений.txt"')
 

    def put_context_in_template(self):
        self.template_docx.render(self.context)

 
    def save_new_files(self, path: str, filename_result: str):
        self.template_docx.save(f'{path}{filename_result}.docx')
 

    def save_new_file(self) -> TextIOBase:
        try:
            # вставить значение из словаря в шаблон word'а
            self.put_context_in_template()
            # укажем путь куда необходимо сохранить новый файл
            path_for_newfile = QtWidgets.QFileDialog.getSaveFileName(
                parent=self,
                caption='Введите куда надо сохранить файл',
                filter='DOCX File (*.docx)'
            )
            filepath = path_for_newfile[0]
            self.template_docx.save(filepath)
            raise EOFError
        except EOFError:
            self.complete_message(message=f'Скрипт завершил свою работу\n')
        except BaseException:
            self.error_message(message='Необходимо указать\n"шаблон файла.docx"')
 
    
    def getting_for_directory(self):
        try:
            path_setting = ((QtWidgets.QFileDialog.getExistingDirectory(
                parent=self,
                caption='Выберите папку, где лежит\n"шаблонизация файлов.xlsx"'
            )).replace('/', '\\')).strip()
            self.download_dict_settings_from_excel(path=path_setting)
            # укажем путь откуда брать шаблон-файл
            template_path = path_setting + '\\шаблоны\\'
            # укажем путь откуда брать файл с заготовкой
            file_path = path_setting + '\\заготовки\\'
            # укажем путь куда необходимо сохранить новый файл
            path_for_newfile = path_setting + '\\итоговые файлы\\'
            for idx in range(len(SETTINGS['Название файла, откуда брать значения'])):
                self.getting_a_template(
                    path=template_path,
                    filename_template=SETTINGS['Название шаблона'][idx]
                )
                self.getting_context_for_insertion(
                    path=file_path,
                    filename_replacement=SETTINGS['Название файла, откуда брать значения'][idx]
                )
                self.put_context_in_template()
                filename = f"{SETTINGS['Название файла, откуда брать значения'][idx]} {SETTINGS['Название итогового файла'][idx]}"
                self.save_new_files(
                    path=path_for_newfile,
                    filename_result=filename
                )
            raise EOFError
        except EOFError:
            self.complete_message(message=f'Скрипт завершил свою работу\n')
        except FileNotFoundError:
            self.error_message(message='Не верно указали папку,\nгде находится файл\n"шаблонизация файлов.xlsx"\nили не указали совсем')
 
 
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    excemple = TemplateWordGenerator()
    excemple.show()
    sys.exit(app.exec_())
